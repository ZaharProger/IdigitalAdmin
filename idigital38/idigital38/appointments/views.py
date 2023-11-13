import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view

from .forms import AppointmentForm
from .models import Appointment


@api_view(['GET', 'POST'])
def appointment_view(request):
    if request.method == 'GET' and request.user.is_authenticated:
        workbook = openpyxl.Workbook()

        workbook.create_sheet('Заявки')
        appointments_sheet = workbook['Заявки']

        appointments_sheet['A1'] = '№'
        appointments_sheet['B1'] = 'ФИО'
        appointments_sheet['C1'] = 'Должность'
        appointments_sheet['D1'] = 'Организация'
        appointments_sheet['E1'] = 'E-mail'
        appointments_sheet['F1'] = 'Номер телефона'
        appointments_sheet['G1'] = 'Формат участия'
        appointments_sheet['H1'] = 'Секции'

        appointments_sheet.column_dimensions['B'].width = 50
        appointments_sheet.column_dimensions['C'].width = 35
        appointments_sheet.column_dimensions['D'].width = 35
        appointments_sheet.column_dimensions['E'].width = 35
        appointments_sheet.column_dimensions['F'].width = 35
        appointments_sheet.column_dimensions['G'].width = 35
        appointments_sheet.column_dimensions['H'].width = 120

        header_font = Font(color='00000000', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for cell in appointments_sheet['1:1']:
            cell.font = header_font
            cell.alignment = header_alignment

        appointments = Appointment.objects.all()
        for i in range(len(appointments)):
            row_number = i + 2
            appointments_sheet.cell(row=row_number, column=1).value = i + 1
            appointments_sheet.cell(row=row_number, column=2).value = appointments[i].name
            appointments_sheet.cell(row=row_number, column=3).value = appointments[i].status
            appointments_sheet.cell(row=row_number, column=4).value = appointments[i].organization
            appointments_sheet.cell(row=row_number, column=5).value = appointments[i].email
            appointments_sheet.cell(row=row_number, column=6).value = appointments[i].phone
            appointments_sheet.cell(row=row_number, column=7).value = appointments[i].participation_type
            appointments_sheet.cell(row=row_number, column=8).value = appointments[i].sections

        workbook.remove(workbook['Sheet'])
        workbook.save('export-data/Idigital38_Reports.xlsx')
        workbook.close()

        with open('export-data/Idigital38_Reports.xlsx', 'rb') as file:
            data = file.read()

        response = HttpResponse(
            data,
            status=status.HTTP_200_OK,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Idigital38_Reports.xlsx"'
    else:
        new_appointment = AppointmentForm(request.data)
        if new_appointment.is_valid():
            new_appointment.save()
            response_status = status.HTTP_200_OK
        else:
            response_status = status.HTTP_400_BAD_REQUEST

        response = Response(
            {'message': ''},
            status=response_status,
            content_type='application/json'
        )

    return response
